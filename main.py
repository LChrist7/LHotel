from datetime import date, timedelta, datetime
from flask import Flask, render_template, url_for, request, g, redirect
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os
import sqlite3
import DBSQL
import string
import sys
import webbrowser
import gclass

DATABASE = '/tmp/hotel.db'
DEBUG = False
SECRET_KEY = 'anua'


def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


app = Flask(__name__, static_url_path="", static_folder=resource_path(
    'static'), template_folder=resource_path("templates"))
app.config.from_object(__name__)
app.config.update(dict(DATABASE=os.path.join(app.root_path, 'hotel.db')))


def connect_db():
    conn = sqlite3.connect(app.config['DATABASE'])
    conn.row_factory = sqlite3.Row
    return conn


def get_db():
    if not hasattr(g, 'link_db'):
        g.link_db = connect_db()
    return g.link_db


@app.route("/search", methods=["POST", "GET"])
def search():
    if request.method == 'POST':
        searchstart = datetime.fromisoformat(request.form['SearchStart'].replace('T', ' '))
        searchend = datetime.fromisoformat(request.form['SearchEnd'].replace('T', ' '))
        db = get_db()
        dbase = DBSQL.DBSQL(db)
        return render_template('search.html', rooms=dbase.makesearch(searchstart, searchend))
    else:
        return render_template('search.html')


@app.route("/cancel", methods=["POST", "GET"])
def cancel():
    if request.method == 'POST' and request.form['numcancel'] != '':
        db = get_db()
        dbase = DBSQL.DBSQL(db)
        dbase.bookcancel(request.form['numcancel'])
        db.commit()
        db.close()
        return redirect("/cancel")
    return render_template('cancel.html')


@app.route("/check", methods=["POST", "GET"])
def check():
    if request.method == 'POST' and "checkstart" in request.form:
        workbook = Workbook()
        sheet = workbook.active
        sheet.freeze_panes = sheet['B2']
        db = get_db()
        dbase = DBSQL.DBSQL(db)
        rooms = dbase.checkbooks()
        for i in range(1, len(rooms) + 1):
            sheet['A' + str(i + 1)] = rooms[i - 1]
        letters = list(string.ascii_uppercase)
        letters.extend([i + b for i in letters for b in letters])
        letters.pop(0)
        start_date = date.fromisoformat(request.form['checkstart'])
        end_date = date.fromisoformat(request.form['checkend'])
        delta = timedelta(days=1)
        alldelta = end_date - start_date
        alldelta = alldelta.days
        for j in letters[0:alldelta + 1]:
            sheet[j + str(1)] = str(start_date)
            sheet.column_dimensions[j].width = 15
            start_date += delta
        for m in range(2, len(rooms) + 2):
            for j in letters[0:alldelta + 1]:
                roomv = sheet['A' + str(m)].value
                datev = sheet[j + str(1)].value
                sqlres = dbase.makecheck(datev, roomv)
                if sqlres and len(sqlres) == 1:
                    red_background = PatternFill(fill_type='solid', fgColor="FFC7CE")
                    sheet[j + str(m)] = sqlres[0].split(" ")[0]
                    sheet[j + str(m)].fill = red_background
                if sqlres and len(sqlres) == 2:
                    sheet.column_dimensions[j].width = 23
                    red_background = PatternFill(fill_type='solid', fgColor="FFC7CE")
                    sheet[j + str(m)] = sqlres[0].split(" ")[0] + ' - ' + sqlres[1].split(" ")[0]
                    sheet[j + str(m)].fill = red_background
        try:
            workbook.save(filename="График.xlsx")
            db.close()
        except PermissionError:
            db.close()
            return '<h2>Файл отчета открыт, чтобы выгрузить новый отчет закройте файл</h2>'
        return redirect('/check')

    if request.method == 'POST' and "allcheckstart" in request.form:
        workbookall = Workbook()
        sheetall = workbookall.active
        db = get_db()
        dbaseall = DBSQL.DBSQL(db)
        row = dbaseall.checkall(request.form['allcheckstart'], request.form['allcheckend'])
        db.close()
        letters = list(string.ascii_uppercase)
        letters.extend([i + b for i in letters for b in letters])
        sheetall.column_dimensions['B'].width = 45
        sheetall.column_dimensions['C'].width = 12
        sheetall.column_dimensions['D'].width = 20
        sheetall.column_dimensions['H'].width = 16
        sheetall.column_dimensions['I'].width = 13
        sheetall.column_dimensions['L'].width = 16
        sheetall.column_dimensions['M'].width = 50
        sheetall['A1'] = '№ заявки'
        sheetall['B1'] = 'ФИО'
        sheetall['C1'] = '№ комнаты'
        sheetall['D1'] = 'Дата рождения'
        sheetall['E1'] = 'Цена'
        sheetall['F1'] = 'Сумма'
        sheetall['G1'] = 'Гостей'
        sheetall['H1'] = 'Полный пансион'
        sheetall['I1'] = 'Полупансион'
        sheetall['J1'] = 'Завтрак'
        sheetall['K1'] = 'От туроператора'
        sheetall['L1'] = 'Трансфер'
        sheetall['M1'] = 'Комментарий'
        i, k = 0, 2
        for item in row:
            sumg = 0
            for j in letters[0:6]:
                sheetall[j + str(k)] = item[i]
                i += 1
            for numguest in range(6, 11):
                if item[numguest]:
                    sumg += 1
            sheetall[letters[6] + str(k)] = sumg
            if item[11] == 0:
                sheetall[letters[7] + str(k)] = ''
            else:
                sheetall[letters[7] + str(k)] = item[11]
            if item[12] == 0:
                sheetall[letters[8] + str(k)] = ''
            else:
                sheetall[letters[8] + str(k)] = item[12]
            if item[13] == 0:
                sheetall[letters[9] + str(k)] = ''
            else:
                sheetall[letters[9] + str(k)] = item[13]
            if item[14] == 1:
                sheetall[letters[10] + str(k)] = 'Да'
            if item[15] == 1:
                sheetall[letters[11] + str(k)] = 'Нужен'
            sheetall[letters[12] + str(k)] = item[16]
            k += 1
            i = 0
        try:
            workbookall.save(filename="Все_бронирования.xlsx")
        except PermissionError:
            return '<h2>Файл отчета открыт, чтобы выгрузить новый отчет закройте файл</h2>'
        return redirect('/check')
    return render_template('check.html')


@app.route("/change", methods=["POST", "GET"])
def change():
    if request.method == 'POST' and 'numchange' in request.form and request.form['numchange']:
        print(request.form)
        db = get_db()
        dbase = DBSQL.DBSQL(db)
        context = list(dbase.viewbook(request.form['numchange']))
        try:
            guests = list(context[0])
        except IndexError:
            return "<h2>Брони с таким номером не существует</h2>"
        info = list(context[1])
        return render_template('change.html', guest1=dict(guests[0]),
                               guest2=(dict(guests[1]) if len(guests) > 1 else []),
                               guest3=(dict(guests[2]) if len(guests) > 2 else []),
                               guest4=(dict(guests[3]) if len(guests) > 3 else []),
                               guest5=(dict(guests[4]) if len(guests) > 4 else []),
                               info=dict(info[0]), sumdiff=int(info[0]['sumbook'] - info[0]['prep']))
    if request.method == 'POST' and 'FullName1' in request.form and request.form['FullName1']:
        db = get_db()
        dbase = DBSQL.DBSQL(db)
        sumbook = (datetime.fromisoformat(request.form['DateEnd'].replace('T', ' ')).date() -
                   datetime.fromisoformat(request.form['DateStart'].replace('T', ' ')).date()).days * int(
            request.form['Price'])
        startdatedef = datetime.fromisoformat(request.form['DateStart'].replace('T', ' '))
        enddatedef = datetime.fromisoformat(request.form['DateEnd'].replace('T', ' '))
        guest1 = gclass.GClass()
        guest2 = gclass.GClass()
        guest3 = gclass.GClass()
        guest4 = gclass.GClass()
        guest5 = gclass.GClass()
        guest1.createguest1(request.form)
        guest2.createguest2(request.form)
        guest3.createguest3(request.form)
        guest4.createguest4(request.form)
        guest5.createguest5(request.form)
        if 'Transfer' in request.form and request.form['Transfer'] == 'on':
            transfer = 1
        else:
            transfer = 0
        if 'Tour' in request.form and request.form['Tour'] == 'on':
            tour = 1
        else:
            tour = 0
        resadd = dbase.updatebook(request.form['numchange2'], guest1, guest2, guest3, guest4, guest5,
                               startdatedef, enddatedef, request.form['Room'],
                               tour, transfer, request.form['Price'], request.form['Prep'],
                               sumbook, str(startdatedef), str(enddatedef), request.form['Comm'])
        db.commit()
        db.close()
        if resadd == 1:
            return '<h2>Бронь не может быть создана. Комната занята в эти даты</h2>'
        if resadd == 0:
            return '<h2>Бронь не может быть создана. Произошла ошибка</h2>'
        return redirect("/change")
    else:
        return render_template('change.html', guest1=[], guest2=[], guest3=[], guest4=[], guest5=[],
                               info=[], sumdiff=0)


@app.route("/", methods=["POST", "GET"])
def index():
    if request.method == 'POST' and 'FullName1' in request.form:
        print(2)
        sumbook = (datetime.fromisoformat(request.form['DateEnd'].replace('T', ' ')).date() -
                   datetime.fromisoformat(request.form['DateStart'].replace('T', ' ')).date()).days * int(
            request.form['Price'])
        startdatedef = datetime.fromisoformat(request.form['DateStart'].replace('T', ' '))
        enddatedef = datetime.fromisoformat(request.form['DateEnd'].replace('T', ' '))
        db = get_db()
        dbase = DBSQL.DBSQL(db)
        guest1 = gclass.GClass()
        guest2 = gclass.GClass()
        guest3 = gclass.GClass()
        guest4 = gclass.GClass()
        guest5 = gclass.GClass()
        guest1.createguest1(request.form)
        guest2.createguest2(request.form)
        guest3.createguest3(request.form)
        guest4.createguest4(request.form)
        guest5.createguest5(request.form)
        if 'Transfer' in request.form and request.form['Transfer'] == 'on':
            transfer = 1
        else:
            transfer = 0
        if 'Tour' in request.form and request.form['Tour'] == 'on':
            tour = 1
        else:
            tour = 0
        print(3)
        resadd = dbase.addbook(request.form['Numbook'], guest1, guest2, guest3, guest4, guest5,
                               startdatedef, enddatedef, request.form['Room'],
                               tour, transfer, request.form['Price'], request.form['Prep'],
                               sumbook, str(startdatedef), str(enddatedef), request.form['Comm'])
        print(4)
        db.commit()
        db.close()
        if resadd == 1:
            return '<h2>Бронь не может быть создана. Комната занята в эти даты</h2>'
        if resadd == 0:
            return '<h2>Бронь не может быть создана. Произошла ошибка</h2>'
        return redirect('/')
    else:
        return render_template('index.html')


if __name__ == '__main__':
    webbrowser.open('http://127.0.0.1:5000/')
    app.run()
